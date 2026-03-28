"""
cortex_document_parser.py — Multi-format Document Parser
=========================================================
Parses PDF, Word, Excel, CSV, and plain text into clean text
for agent consumption or SurfSense ingestion.

Size routing:
  < 8000 tokens (~6000 words) → returns text directly (inject as context)
  ≥ 8000 tokens               → chunks text → returns list for SurfSense push

Supported formats:
  .pdf   → PyMuPDF (fitz)
  .docx  → python-docx
  .xlsx  → openpyxl
  .csv   → built-in csv module
  .txt / .md → built-in open()

Requires (optional — only needed for respective formats):
  pip install PyMuPDF python-docx openpyxl
"""

import csv
import io
import os
from typing import Union


# Approximate token count threshold (1 token ≈ 0.75 words)
_TOKEN_THRESHOLD = 8_000
_CHARS_PER_TOKEN = 4          # conservative estimate
_CHUNK_SIZE_CHARS = 24_000    # ~6000 tokens per chunk for SurfSense


class ParseError(Exception):
    """Raised when a document cannot be parsed."""


class ParseResult:
    """
    Result of document parsing.

    Attributes:
        text:       Full extracted text (always populated).
        chunks:     Non-empty list means document is large → use SurfSense push.
        filename:   Original filename.
        mime_type:  Detected MIME type.
        page_count: Number of pages (PDF) or sheets (Excel), else None.
    """

    def __init__(
        self,
        text: str,
        filename: str,
        mime_type: str,
        page_count: int | None = None,
    ):
        self.text = text
        self.filename = filename
        self.mime_type = mime_type
        self.page_count = page_count
        self.chunks: list[str] = []

        # Auto-chunk if above threshold
        if len(text) >= _CHUNK_SIZE_CHARS:
            self.chunks = _chunk_text(text, _CHUNK_SIZE_CHARS)

    @property
    def is_large(self) -> bool:
        return bool(self.chunks)

    @property
    def token_estimate(self) -> int:
        return len(self.text) // _CHARS_PER_TOKEN


class CortexDocumentParser:
    """
    Parse documents from bytes into clean text.

    Usage:
        parser = CortexDocumentParser()
        result = await parser.parse(file_bytes, filename="report.pdf")
        if result.is_large:
            # push result.chunks to SurfSense
        else:
            # inject result.text as context
    """

    async def parse(self, file_bytes: bytes, filename: str) -> ParseResult:
        """
        Parse a document.

        Args:
            file_bytes: Raw file bytes.
            filename:   Original filename (used to detect format).

        Returns:
            ParseResult with extracted text and chunking info.

        Raises:
            ParseError if the format is unsupported or parsing fails.
        """
        import asyncio
        import functools

        ext = _get_ext(filename)
        parser_fn = _PARSERS.get(ext)
        if not parser_fn:
            raise ParseError(
                f"Unsupported format: {ext!r}. "
                f"Supported: {', '.join(_PARSERS.keys())}"
            )

        loop = asyncio.get_event_loop()
        text, page_count = await loop.run_in_executor(
            None,
            functools.partial(parser_fn, file_bytes),
        )

        mime = _MIME_TYPES.get(ext, "application/octet-stream")
        return ParseResult(
            text=text.strip(),
            filename=filename,
            mime_type=mime,
            page_count=page_count,
        )

    def supports(self, filename: str) -> bool:
        return _get_ext(filename) in _PARSERS


# ------------------------------------------------------------------
# Format-specific parsers (all sync — run in executor)
# ------------------------------------------------------------------

def _parse_pdf(file_bytes: bytes) -> tuple[str, int]:
    try:
        import fitz  # PyMuPDF
    except ImportError as e:
        raise ParseError("PyMuPDF not installed. Run: pip install PyMuPDF") from e

    doc = fitz.open(stream=file_bytes, filetype="pdf")
    pages = []
    for page in doc:
        pages.append(page.get_text("text"))
    doc.close()
    return "\n\n".join(pages), len(pages)


def _parse_docx(file_bytes: bytes) -> tuple[str, int]:
    try:
        from docx import Document
    except ImportError as e:
        raise ParseError("python-docx not installed. Run: pip install python-docx") from e

    doc = Document(io.BytesIO(file_bytes))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs), None


def _parse_xlsx(file_bytes: bytes) -> tuple[str, int]:
    try:
        import openpyxl
    except ImportError as e:
        raise ParseError("openpyxl not installed. Run: pip install openpyxl") from e

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheets_text = []
    for sheet in wb.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append("\t".join(cells))
        if rows:
            sheets_text.append(f"## Sheet: {sheet.title}\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(sheets_text), len(wb.sheetnames) if sheets_text else 0


def _parse_csv(file_bytes: bytes) -> tuple[str, int]:
    text = file_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = ["\t".join(row) for row in reader if any(c.strip() for c in row)]
    return "\n".join(rows), None


def _parse_text(file_bytes: bytes) -> tuple[str, int]:
    return file_bytes.decode("utf-8", errors="replace"), None


# ------------------------------------------------------------------
# Registry
# ------------------------------------------------------------------

_PARSERS = {
    ".pdf":  _parse_pdf,
    ".docx": _parse_docx,
    ".doc":  _parse_docx,   # older Word; python-docx handles most .doc files
    ".xlsx": _parse_xlsx,
    ".xls":  _parse_xlsx,   # openpyxl reads most .xls via compatibility
    ".csv":  _parse_csv,
    ".txt":  _parse_text,
    ".md":   _parse_text,
}

_MIME_TYPES = {
    ".pdf":  "application/pdf",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".doc":  "application/msword",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls":  "application/vnd.ms-excel",
    ".csv":  "text/csv",
    ".txt":  "text/plain",
    ".md":   "text/markdown",
}


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------

def _get_ext(filename: str) -> str:
    _, ext = os.path.splitext(filename.lower())
    return ext


def _chunk_text(text: str, chunk_size: int) -> list[str]:
    """Split text into chunks at paragraph boundaries where possible."""
    paragraphs = text.split("\n\n")
    chunks = []
    current = []
    current_len = 0

    for para in paragraphs:
        para_len = len(para)
        if current_len + para_len > chunk_size and current:
            chunks.append("\n\n".join(current))
            current = []
            current_len = 0
        # If a single paragraph is larger than chunk_size, force-split it
        if para_len > chunk_size:
            for i in range(0, para_len, chunk_size):
                chunks.append(para[i:i + chunk_size])
        else:
            current.append(para)
            current_len += para_len + 2  # +2 for \n\n

    if current:
        chunks.append("\n\n".join(current))

    return [c for c in chunks if c.strip()]
